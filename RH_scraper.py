from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font
import openpyxl
import pandas as pd
from datetime import datetime, date

def date_str2obj(date_str):
    tday = str(date.today().year)
    cal_dict = {"Jan" : "1", "Feb" : "2", "Mar" : "3", "Apr" : "4", "May" : "5", "Jun" : "6", "Jul" : "7", "Aug" : "8", "Sep" : "9", "Oct" : "10", "Nov" : "11", "Dec" : "12"}
    month_num = cal_dict[(date_str[0:3])]
    date_num = date_str[4:6]
    dd = month_num +'/'+ date_num + '/' + (tday[2:4])
    dt_obj = datetime.strptime(dd, '%m/%d/%y')
    return dt_obj

# #######            WEB SCRAPPER CODE                            ##################################################################

# #URLS and Xpaths
redhat_home = 'https://ul2647.principal.com/users/login'
redhat_errata = 'https://ul2647.principal.com/errata?page=1&per_page=100&sortBy=updated&sortOrder=DESC&search=type%3D%20security'

#Xpath to get All cves from Advisory page
num_cve = '//span[@ng-repeat="cve in errata.cves"]'
# Xpath to find all rows in errata page
num_rows = '//tr[@row-select="errata"]/td/a[@ui-sref="erratum.info({errataId: errata.id})"]'
#Xpath to find all Cves present on Advisory Page
cve_xpath = '//span[@ng-repeat="cve in errata.cves"]/child::a'
#Xpath to find the type of alert
alert_type = '//header[@data-block="header"]'
#Xpath to find the Severity of Alert
alert_sev = '//dt[contains(text(), "Severity")]//following::dd'
#Xpath to find OS
alert_OS = '//p[@class="info-paragraph ng-binding"][1]'
#Xpath to find Description
alert_desc = '//p[@class="info-paragraph ng-binding"][2]'

driver = webdriver.Chrome()
#driver = webdriver.Chrome("C:/Work/patch_automation/code_files/web_scraping")
driver.get(redhat_home)
time.sleep(2)
ele_login = driver.find_element_by_id("login_login")
ele_pwd = driver.find_element_by_id("login_password")
ele_submit = driver.find_element_by_id("login_submit_btn")

#clear any pre-populated text in the form field
ele_login.clear()
ele_pwd.clear()
ele_login.send_keys("T429016")
ele_pwd.send_keys("todomeda@21")

ele_submit.send_keys(Keys.RETURN) 

driver.get(redhat_errata)
time.sleep(10)
checkbox = driver.find_element_by_xpath('//input[@ng-model="showApplicable"]')
checkbox.click()
time.sleep(10)
id_list = []
date_list = []
href_list= []

row_count = len(driver.find_elements_by_xpath(num_rows))
print('number of rows are: ' + str(row_count))

print("\n >> Pulling all the info from main table")
table = driver.find_element_by_xpath('//table[@class="table table-striped table-bordered"]')
for row in table.find_elements_by_xpath('.//tr'):
    for td in row.find_elements_by_xpath('.//a[@ui-sref="erratum.info({errataId: errata.id})"]'):
        #print(td.text)
        id_list.append(td.text)

    for dd in row.find_elements_by_xpath('.//span/parent::span/parent::short-date-time[@date="errata.updated"]'):
        #print(dd.text)
        date_obj = date_str2obj(dd.text)
        date_list.append(date_obj)

    for lk in row.find_elements_by_xpath('.//a[@ui-sref="erratum.info({errataId: errata.id})"]'):
        #print(lk.get_attribute('href'))
        href_list.append(lk.get_attribute('href'))
#print(id_list)         
#print(date_list)
#print(href_list)
print('>> Initial Dataframe created')
print('>> Enter Date range to filter products\n    Syntax for Date should be exactly as follows\n    Example Date: September 20 => Sep 20')
df= pd.DataFrame(list(zip(id_list, href_list, date_list)),columns = ['Advisory','Link', 'Date'])
from_date = date_str2obj(input('\n>>  Enter Starting/FROM Date: '))
to_date = date_str2obj(input('\n>>  Enter End/TO Date: '))
after_from_date = df['Date'] >= from_date
before_to_date = df['Date'] <= to_date
between_dates = after_from_date & before_to_date
df = df.loc[between_dates]
# Convert 'Date' column Objects to string
df['Date'] = df['Date'].astype(str)
df = df[df['Advisory'].str.contains('RHSA')]
print('Filtered DataFrame')
print(df)

print(">>  Writing Filtered Dataframe to file")
with pd.ExcelWriter('df_test.xlsx') as writer:
    df.to_excel(writer, index=False)
print(">> df_test.xlsx" + " saved to " + os.getcwd())   

newwb = openpyxl.load_workbook('df_test.xlsx')    
ws = newwb.active
ws['D1'].value = 'Type'
ws['D1'].font = Font(bold=True)
ws['E1'].value = 'Source Rating'
ws['E1'].font = Font(bold=True)
ws['F1'].value = 'CVE'
ws['F1'].font = Font(bold=True)
ws['G1'].value = 'Description'
ws['G1'].font = Font(bold=True)
ws['H1'].value = 'OS'
ws['H1'].font = Font(bold=True)

for i in range (1, ws.max_row +1):
    cve_list = []
    link = ws['B' + str(i+1)].value
    if link == None:
        print("None detected")
    else:
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[i])
        print('>>  opening_link: ' + link)
        driver.get(link)
        #time.sleep(10)
        try:
            new_page = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, num_cve)))
        
        except Exception as e:
            print(e)
        finally:
            atype = new_page.find_element_by_xpath(alert_type)
            atype = atype.text
            atype = atype.partition(': ')[2]
            ws['D' + str(i+1)].value = atype
            print('Type: ' + atype + '\n')
            sev = new_page.find_element_by_xpath(alert_sev)
            print('Severity: ' + sev.text + '\n')
            ws['E' + str(i+1)].value = sev.text
            desc = new_page.find_element_by_xpath(alert_desc)
            print('Description: ' + desc.text + '\n')
            ws['G' + str(i+1)].value = desc.text
            OS = new_page.find_element_by_xpath(alert_OS)
            print('OS: ' + OS.text + '\n')
            ws['H' + str(i+1)].value = OS.text 
            cves = new_page.find_elements_by_xpath(cve_xpath)
            print(len(cves))
            for cve in cves:
                print(cve)
                cve_list.append(cve.text)
        ws['F' + str(i+1)].value = ','.join(map(str, cve_list))

#WRITE TO EXCEL FILE ###########
newwb.save('Redhat Updates.xlsx')
print(">> Excel sheet prepared.")
print(">> Redhat Updates.xlsx" + " saved to " + os.getcwd())
# CLOSE BROWSER
driver.quit()
