# This code is used to pull comma separated CVEs from redhat advisory and place them in new line so that they can be further
# used by VulnDB Script. It creates two new sheets in the workbook to do that.

import openpyxl
from openpyxl.styles import Font
import pandas as pd
import os

newwb = openpyxl.load_workbook('Redhat Updates.xlsx')    
newwb.create_sheet(title='Security Updates', index=2)
newwb.create_sheet(title='Master CVE', index=3)

sheet1 = newwb['Sheet1']
sheet2 = newwb['Security Updates']
sheet3 = newwb['Master CVE']

sheet2['A1'].value = 'Advisory'
sheet2['A1'].font = Font(bold=True)
sheet2['B1'].value = 'Details'
sheet2['B1'].font = Font(bold=True)

sheet3['A1'].value = 'Details'
sheet3['A1'].font = Font(bold=True)
sheet3['B1'].value = 'Score'
sheet3['B1'].font = Font(bold=True)
sheet3['C1'].value = 'Location'
sheet3['C1'].font = Font(bold=True)
sheet3['D1'].value = 'Exploit'
sheet3['D1'].font = Font(bold=True)
sheet3['E1'].value = 'Description'
sheet3['E1'].font = Font(bold=True)

n = 1
X=1
h=n
for i in range(1, sheet1.max_row + 1):
    sheet2['A' + str(X+1)].value = sheet1['A' + str(i+1)].value
    CVE_String = sheet1['F' + str(i + 1)].value
    if CVE_String is not None:
        cve_list = CVE_String.split(",")
        l = len(cve_list)
        #print('length of the list is '+ str(l))
        
        for z in cve_list:
            sheet2['B' + str(h+1)].value = z
            sheet3['A' + str(n+1)].value = z
            n=n+1
            h=h+1
        h=h+1
        X=h
newwb.save('Redhat Updates.xlsx')

# Pandas to the rescue
df1 = pd.read_excel('Redhat Updates.xlsx', sheet_name="Sheet1")
df2 = pd.read_excel('Redhat Updates.xlsx', sheet_name="Security Updates")
df3 = pd.read_excel('Redhat Updates.xlsx', sheet_name="Master CVE")
# remove duplicates
df3 = (df3[~df3.duplicated(subset=['Details'])])

# sort the values
df3 = df3.sort_values('Details')

print(">> total unique CVEs: " + str(len(df3)))

# write all the sheets to single file
with pd.ExcelWriter('Redhat Updates.xlsx') as writer:
    df1.to_excel(writer, sheet_name='Sheet1', index=False)
    df2.to_excel(writer, sheet_name='Security Updates', index=False)
    df3.to_excel(writer, sheet_name='Master CVE', index=False)

print(">> Excel sheet prepared.")
print(">> Redhat Updates.xlsx" + " saved to " + os.getcwd())
